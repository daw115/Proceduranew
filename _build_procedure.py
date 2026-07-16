#!/usr/bin/env python3
"""Buduje rozszerzoną PROCEDURĘ IDCC TSO v5.2 z pełnym katalogiem CCM monitoring.

Wejście:
  /Volumes/SSD/Downloads/PROCEDURA_IDCC_TSO_v5_1_v9.html  (baza)
  popup_content_IDCC.json                                  (126 FID-ów)
  popup_states_IDCC.json                                   (82 stany)
  Pulpit_IDCC_CCM_opisy_kafelkow.xlsx                      (kafelki, legenda, backup)

Wyjście:
  PROCEDURA_IDCC_TSO_v5_2.html                             (bez edytora)
  PROCEDURA_IDCC_TSO_v5_2_EDYTOR.html                      (z edytorem)
"""
import json, os, re, html
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
# Plik bazowy v5.1 nie jest częścią repo — ścieżkę podaje się przez PROCEDURA_V51_SRC
SRC  = Path(os.environ.get('PROCEDURA_V51_SRC',
                           '/Volumes/SSD/Downloads/PROCEDURA_IDCC_TSO_v5_1_v9.html'))
if not SRC.exists():
    raise SystemExit(
        f"Brak pliku bazowego v5.1: {SRC}\n"
        "Ustaw zmienną środowiskową PROCEDURA_V51_SRC na ścieżkę do "
        "PROCEDURA_IDCC_TSO_v5_1_v9.html (plik spoza repozytorium).")

def esc(s):
    if s is None: return ''
    return html.escape(str(s))

# ── 1. Wczytaj źródła ──────────────────────────────────────────────────────
with open(ROOT/'popup_content_IDCC.json', encoding='utf-8') as f:
    popups = json.load(f)
with open(ROOT/'popup_states_IDCC.json', encoding='utf-8') as f:
    states = json.load(f)
wb = load_workbook(ROOT/'Pulpit_IDCC_CCM_opisy_kafelkow.xlsx', data_only=True)

# Spis kafelków
spis_rows = []
for r in wb['Spis kafelków'].iter_rows(min_row=2, values_only=True):
    if r and r[0]:
        spis_rows.append(r)

# Legenda
legenda_rows = []
for r in wb['Legenda kolorów'].iter_rows(min_row=2, values_only=True):
    if r and r[0]: legenda_rows.append(r)

# Backup
backup_rows = []
for r in wb['Backup - zasady HTML'].iter_rows(min_row=2, values_only=True):
    if r and r[0]: backup_rows.append(r)

# Stany kafelków per plik
stany_rows = []
for r in wb['Kafelki - stany'].iter_rows(min_row=2, values_only=True):
    if r and r[0]: stany_rows.append(r)

# ── 2. Generuj sekcję 7.9 — Pełny katalog kafelków ─────────────────────────
def color_dot(name, size=20):
    """Mapuje nazwę koloru na ikonę wyciętą z legendy CCM (img_ccm/status/)."""
    n = (name or '').lower()
    icon = None
    if 'puls' in n:                                  icon = 'ziel_puls'
    elif 'ciemnozielon' in n and 'po cet' in n:      icon = 'ciemnoziel_W'
    elif 'ciemnozielon' in n:                        icon = 'ziel_W'
    elif 'zielon' in n and ("'r'" in n or n.endswith(' r') or ' r ' in n+' '): icon = 'ziel_R'
    elif 'zielon' in n:                              icon = 'ziel'
    elif 'czerwon' in n and '?' in n:                icon = 'czerw_q'
    elif 'czerwon' in n and '!' in n:                icon = 'czerw_excl'
    elif 'czerwon' in n:                             icon = 'czerw'
    elif 'pomarań' in n:                             icon = 'pomar'
    elif 'czarny' in n:                              icon = 'czarny'
    elif 'szary' in n and '?' in n:                  icon = 'szary'
    elif 'szary' in n:                               icon = 'szary_jasny'
    elif 'fiolet' in n:                              icon = 'fiolet'
    elif 'puste' in n or 'białe' in n or 'biale' in n: icon = 'szary_jasny'
    elif 'kółko' in n or 'ładow' in n:
        return '<span class="cdot-spin"></span>'
    if icon is None:
        icon = 'szary_jasny'
    h = round(size * 0.88)
    return f'<img src="img_ccm/status/{icon}.png" class="sdot" width="{size}" height="{h}" alt="{esc(name)}">'

def make_pelna_legenda():
    rows = []
    for kolor, opis in legenda_rows:
        rows.append(f'<tr><td style="white-space:nowrap">{color_dot(kolor, size=26)} <strong>{esc(kolor)}</strong></td><td>{esc(opis)}</td></tr>')
    return '\n'.join(rows)

def make_katalog_kafelkow():
    """Buduje tabelę 73 kafelków z TET/CET/profilem."""
    by_phase = {}
    for r in spis_rows:
        phase, name, fid, tet, cet, profil, n_states = r[:7]
        by_phase.setdefault(phase, []).append((name, fid, tet, cet, profil, n_states))
    colgroup = ('<colgroup><col style="width:30%"><col style="width:13%"><col style="width:8%">'
                '<col style="width:8%"><col style="width:33%"><col style="width:8%"></colgroup>')
    out = []
    for phase, items in by_phase.items():
        out.append(f'<h5 class="subhead">{esc(phase)}</h5>')
        out.append(f'<div class="tbl-wrap"><table class="tile-cat">{colgroup}')
        out.append('<thead><tr><th>Kafelek</th><th>FID</th><th>TET</th><th>CET</th><th>Profil statusów</th><th style="text-align:center">Stanów</th></tr></thead><tbody>')
        for name, fid, tet, cet, profil, n in items:
            out.append(f'<tr><td>{esc(name)}</td><td><code>{esc(fid)}</code></td><td>{esc(tet)}</td><td>{esc(cet)}</td><td>{esc(profil)}</td><td style="text-align:center">{esc(n)}</td></tr>')
        out.append('</tbody></table></div>')
    return '\n'.join(out)

def make_katalog_FID():
    """126 pozycji popup_content — pełne opisy plików."""
    by_phase = {}
    for it in popups:
        by_phase.setdefault(it['obszar'], []).append(it)
    out = []
    for phase, items in by_phase.items():
        anchor = re.sub(r'[^a-z0-9]+','-', phase.lower()).strip('-')
        out.append(f'<h5 class="subhead" id="cat-{anchor}">{esc(phase)}</h5>')
        for it in items:
            out.append(f'''<div class="fid-card">
  <div class="fid-head"><span class="fid-code">{esc(it.get("kod",""))}</span><span class="fid-name">{esc(it.get("nazwa_przeplywu",""))}</span><span class="fid-meta">TET <strong>{esc(it.get("target_end_time",""))}</strong> &nbsp;|&nbsp; CET <strong>{esc(it.get("critical_end_time",""))}</strong> &nbsp;|&nbsp; profil {esc(it.get("profil",""))}</span></div>
  <div class="fid-body">
    <p><strong>Opis:</strong> {esc(it.get("opis_pliku",""))}</p>
    <p><strong>Ścieżka pliku:</strong> {esc(it.get("sciezka_pliku",""))}</p>
    <p><strong>Źródło:</strong> {esc(it.get("zrodlo",""))} &nbsp;|&nbsp; <strong>Kontakt:</strong> {esc(it.get("kontakt",""))}</p>
    {('<p class="uw"><strong>Uwagi:</strong> '+esc(it.get("uwagi",""))+'</p>') if it.get('uwagi') else ''}
  </div>
</div>''')
    return '\n'.join(out)

def make_katalog_stanow():
    """82 warianty stanów per profil — z opisem i procedurą."""
    by_var = {}
    for s in states:
        by_var.setdefault(s['wariant'], []).append(s)
    out = []
    for var, items in by_var.items():
        var_anchor = re.sub(r'[^a-z0-9]+','-', var.lower()).strip('-')
        out.append(f'<h5 class="subhead" id="st-{var_anchor}">{esc(var)}</h5>')
        colgroup = ('<colgroup><col style="width:32px"><col style="width:86px"><col style="width:190px">'
                    '<col style="width:30%"><col></colgroup>')
        out.append(f'<div class="tbl-wrap"><table class="state-cat">{colgroup}')
        out.append('<thead><tr><th>#</th><th>Kat.</th><th>Stan kafelków</th><th>Opis</th><th>Działanie</th></tr></thead><tbody>')
        for s in items:
            k1 = s.get('kafelek1','') or ''
            k2 = s.get('kafelek2','') or ''
            # etykiety kolumn z profilu wariantu (np. "CCtool wysł." / "CCtool zwalid.")
            stan_str = f'<div class="tile-state">{color_dot(k1)}<span>{esc(k1)}</span></div>'
            if k2: stan_str += f'<div class="tile-state">{color_dot(k2)}<span>{esc(k2)}</span></div>'
            kat_clr = {'A':'#2e7d32','B':'#f57c00','C':'#c62828','D':'#7b1fa2'}.get(str(s.get('kategoria',''))[:1], '#595959')
            out.append(f'<tr><td>{esc(s.get("nr",""))}</td><td><span class="kat-badge" style="background:{kat_clr}">{esc(s.get("kategoria",""))}</span> {esc(s.get("kategoria_nazwa",""))}</td><td>{stan_str}</td><td>{esc(s.get("opis_stanu",""))}</td><td>{s.get("procedura_html","")}</td></tr>')
        out.append('</tbody></table></div>')
    return '\n'.join(out)

def make_backup_scen():
    out = ['<div class="tbl-wrap"><table>']
    out.append('<thead><tr><th style="width:80px">Kod</th><th style="width:30%">Tytuł</th><th>Opis / działanie</th></tr></thead><tbody>')
    for kod, tyt, opis in backup_rows:
        out.append(f'<tr><td><code>{esc(kod)}</code></td><td><strong>{esc(tyt)}</strong></td><td>{esc(opis)}</td></tr>')
    out.append('</tbody></table></div>')
    return '\n'.join(out)

# ── 3. Buduj cały nowy blok 7.9–7.15 ──────────────────────────────────────
CCM_EXT = f'''
    <!-- ═══ ROZSZERZENIE CCM — pełny katalog monitoringu ═══ -->
    <hr class="section-divider">
    <h2 id="s7m" class="anchor-target">7M. Pełny katalog monitoringu CCM (rozszerzenie)</h2>

    <div class="box box-info">
      <strong>ℹ Cel rozszerzenia</strong>
      Sekcja 7M stanowi rozszerzenie sekcji 7. Zawiera kompletny katalog 73 kafelków pulpitu dyspozytorskiego CCM dla wszystkich iteracji IDCC (a/b/c/d + IDA3), pełną legendę 14 statusów kolorystycznych, katalog 126 plików FIDx (opis, ścieżka, źródło, kontakt) oraz katalog 82 wariantów stanów kafelków per profil statusów z opisem działania. Materiał oparty na dokumentach: CCM Instrukcja Użytkownika Dyspozytora ID v2.5, popup_content_IDCC, popup_states_IDCC, Pulpit_IDCC_CCM_opisy_kafelkow.
    </div>

    <h3 id="s7m1" class="anchor-target">7M.1. Tryby odświeżania i konfiguracja widoku</h3>
    <div class="tbl-wrap"><table>
      <thead><tr><th>Tryb</th><th>Częstotliwość</th><th>Zastosowanie</th></tr></thead>
      <tbody>
        <tr><td><strong>Ultraszybki</strong></td><td>10 s</td><td>Aktywne okno krytyczne (30 min przed/po CET); monitoring zmian statusów w czasie rzeczywistym.</td></tr>
        <tr><td><strong>Szybki</strong></td><td>30 s</td><td>Tryb domyślny w trakcie iteracji IDCC; równowaga między aktualnością a obciążeniem przeglądarki.</td></tr>
        <tr><td><strong>Wolny</strong></td><td>5 min</td><td>Okres między iteracjami; oszczędność zasobów stacji dyspozytorskiej.</td></tr>
      </tbody>
    </table></div>

    <div class="box box-warn">
      <strong>⚠ Konfiguracja pulpitu</strong>
      <ul>
        <li><strong>Ikona koła zębatego</strong> (górny prawy róg) — dostęp do konfiguracji widoku: ukrywanie/pokazywanie kolumn, kolejność wierszy, motyw kolorystyczny, wielkość czcionki.</li>
        <li><strong>Ikona kłódki</strong> — blokada widoku przed przypadkowym kliknięciem; włączać przy długich oknach monitoringu bez planowanej interakcji.</li>
        <li><strong>Przycisk „Test powiadomień"</strong> — weryfikacja działania alertów dźwiękowych i wizualnych przed rozpoczęciem dyżuru.</li>
        <li><strong>Lista pulpitów</strong> (menu boczne) — szybkie przełączanie między pulpitem IDCC, IDCC z ATC, Kreatorem IDCF i pulpitami pomocniczymi.</li>
      </ul>
    </div>

    <h3 id="s7m2" class="anchor-target">7M.2. Pełna legenda statusów kolorystycznych (14 kategorii)</h3>
    <div class="tbl-wrap"><table>
      <thead><tr><th style="width:25%">Status</th><th>Znaczenie</th></tr></thead>
      <tbody>
        {make_pelna_legenda()}
      </tbody>
    </table></div>

    <h3 id="s7m3" class="anchor-target">7M.3. Profile statusów</h3>
    <p>Każdy kafelek w CCM jest przypisany do jednego z pięciu profili, które definiują liczbę i znaczenie kolumn statusu w wierszu monitoringu:</p>
    <div class="tbl-wrap"><table>
      <thead><tr><th>Profil</th><th>Kolumny statusu</th><th>Zastosowanie</th></tr></thead>
      <tbody>
        <tr><td><strong>2CCT</strong></td><td>CCTool wysł. + CCTool zwalid.</td><td>Pliki przekazywane do Core CC Tool z oczekiwanym ACK (np. FIDx-607, -617, -710, -831, -928).</td></tr>
        <tr><td><strong>sFTP/ZP</strong></td><td>sFTP/ZP wysł.</td><td>Pliki wysyłane przez SFTP/ZP do Perun4V (np. PERUN_IDx, raporty RAP_PERUN, FIDx-921).</td></tr>
        <tr><td><strong>MinIO</strong></td><td>MinIO</td><td>Składowe paczek FBA/ATC i wyniki pośrednie monitorowane przez dostępność w repozytorium MinIO.</td></tr>
        <tr><td><strong>MinIO + sFTP/ZP</strong></td><td>MinIO + sFTP/ZP wysł.</td><td>Agregaty paczek z monitoringiem zarówno dostępności składowych, jak i potwierdzenia wysyłki paczki.</td></tr>
        <tr><td><strong>2CCT + MinIO</strong></td><td>CCTool wysł. + CCTool zwalid. + MinIO</td><td>Pliki wyników FBA (np. FIDx-710) — wymagają potwierdzenia w CCCt i są równolegle dostępne w MinIO.</td></tr>
      </tbody>
    </table></div>

    <h3 id="s7m4" class="anchor-target">7M.4. Pełny katalog kafelków (73 pozycje)</h3>
    <p>Katalog wszystkich kafelków dostępnych na pulpicie dyspozytorskim CCM dla procesu IDCC, w podziale na iteracje i fazy procesu. Wartości TET (Target End Time) i CET (Critical End Time) podane w czasie lokalnym D / D-1 zgodnie z harmonogramem PSE.</p>
    {make_katalog_kafelkow()}

    <h3 id="s7m5" class="anchor-target">7M.5. Katalog plików FIDx — opisy, ścieżki, źródła (126 pozycji)</h3>
    <p>Pełen katalog plików monitorowanych w CCM w procesie IDCC. Każdy wpis zawiera: kod FIDx, nazwę przepływu, TET/CET, profil statusu, opis biznesowy, ścieżkę przekazania (system → system) oraz źródło i kontakt operacyjny.</p>
    {make_katalog_FID()}

    <h3 id="s7m6" class="anchor-target">7M.6. Katalog stanów kafelków per profil (82 warianty)</h3>
    <p>Każdy profil statusów ma zdefiniowany pełen katalog możliwych kombinacji stanów kafelków, opis sytuacji i procedurę działania dyspozytora. Kategoria A — OK (brak działań), B — obserwacja/przygotowanie, C — wymagana interwencja, D — eskalacja.</p>
    {make_katalog_stanow()}

    <h3 id="s7m7" class="anchor-target">7M.7. Scenariusze backupowe i działania (S.1–S.13, R01–R20)</h3>
    <p>Mapowanie skróconych kodów scenariuszy używanych w katalogu stanów (kolumna „Działanie") na pełne opisy postępowania. Dla pełnych kart ryzyk patrz rozdział 9.</p>
    {make_backup_scen()}

    <h3 id="s7m8" class="anchor-target">7M.8. Kody ACK i błędy walidacyjne CCM</h3>
    <div class="tbl-wrap"><table>
      <thead><tr><th style="width:60px">Kod</th><th>Komunikat</th><th>Działanie dyspozytora</th></tr></thead>
      <tbody>
        <tr><td><code>00</code></td><td>OK / Processed</td><td>Brak działań — plik przyjęty i zwalidowany.</td></tr>
        <tr><td><code>20</code></td><td>gateClosed</td><td>Bramka zamknięta — skontaktować się z administratorem CCCt; nie ponawiać wysyłki bez potwierdzenia.</td></tr>
        <tr><td><code>21</code></td><td>duplicatedVersion</td><td>Wersja pliku już istnieje — podbić numer wersji i ponowić wysyłkę.</td></tr>
        <tr><td><code>22</code></td><td>Negative constraint value</td><td>Wartość ograniczenia ujemna — poprawić dane wejściowe i ponowić.</td></tr>
        <tr><td><code>23</code></td><td>Invalid XML/schema</td><td>Niezgodność schematu — pobrać aktualny szablon, przegenerować plik, ponowić.</td></tr>
        <tr><td><code>24</code></td><td>Missing mandatory field</td><td>Brak pola obowiązkowego — uzupełnić i ponowić.</td></tr>
        <tr><td><code>25</code></td><td>Time series out of range</td><td>Zakres czasowy poza dobą biznesową — zweryfikować datę i zakres TS.</td></tr>
        <tr><td><code>30</code></td><td>Authorization failed</td><td>Brak uprawnień — sprawdzić certyfikat / konto serwisowe, eskalować do CCC.</td></tr>
        <tr><td><code>40</code></td><td>System error</td><td>Błąd techniczny CCCt — odnotować, eskalować do CCC + PSE Innowacje.</td></tr>
      </tbody>
    </table></div>

    <h3 id="s7m9" class="anchor-target">7M.9. Powiadomienia i alerty CCM</h3>
    <ul>
      <li><strong>Alert wizualny</strong> — kafelek przechodzi w stan pulsacji (np. nowa wersja IVA BACKUP); wymaga potwierdzenia operatora.</li>
      <li><strong>Alert dźwiękowy</strong> — w trybie ultraszybkim (10 s) sygnalizuje zmianę statusu z zielonego na inny.</li>
      <li><strong>Pop-up systemowy</strong> — przy 30 min, 15 min i 5 min przed CET dla kafelków pozostających w stanie szarym/pomarańczowym.</li>
      <li><strong>Lista zdarzeń</strong> (panel boczny) — log wszystkich zmian statusów w bieżącej dobie; eksport do CSV przyciskiem w prawym górnym rogu.</li>
    </ul>

    <h3 id="s7m10" class="anchor-target">7M.10. Eksport danych i ślad operacyjny</h3>
    <div class="box box-info">
      <strong>ℹ Eksport monitoringu</strong>
      <ul>
        <li><strong>Eksport stanu pulpitu</strong> — przycisk „Eksportuj" → CSV lub PDF; zawiera snapshot wszystkich kafelków z timestampem.</li>
        <li><strong>Eksport logu zdarzeń</strong> — pełna historia zmian statusów w bieżącej dobie biznesowej.</li>
        <li><strong>Eksport raportu dyżurowego</strong> — zestawienie ręcznych interwencji (wysyłki przed/po CET, statusy nadane ręcznie „R", scenariusze S.x i ryzyka R.x).</li>
      </ul>
      Wszystkie eksporty zachowywać razem z dokumentacją dyżurową przez minimum 90 dni (zgodnie z wymogami audytu CCC).
    </div>

    <h3 id="s7m11" class="anchor-target">7M.11. Pulpit IDCF (Kreator) — różnice względem pulpitu IDCC</h3>
    <p>Pulpit dyspozytorski Kreatora IDCF (GLSK / CBCORA) jest osobnym widokiem CCM dostępnym z menu „Lista pulpitów". Monitoruje on przygotowanie i publikację plików FIDx-607 (GLSK) i FIDx-617 (CBCORA) na potrzeby iteracji IDCC(b/c/d).</p>
    <ul>
      <li>Sekcja <strong>„Generowanie GLSK"</strong> — status uruchomienia kreatora, dostępność plików wejściowych KreatorDACF_CGMES, wynik replacement strategy.</li>
      <li>Sekcja <strong>„Generowanie CBCORA"</strong> — status filtra CB, kompletność CNEC, wersjonowanie pliku CB.</li>
      <li>Sekcja <strong>„Publikacja"</strong> — status wysyłki przez Connector 2.0 do CCM i CCCt (profil 2CCT).</li>
      <li>Pełną procedurę obsługi Kreatora IDCF zawiera <a href="Procedura_KreatorIDCF_GLSK_CBCORA.html">Procedura_KreatorIDCF_GLSK_CBCORA.html</a>.</li>
    </ul>

    <h3 id="s7m12" class="anchor-target">7M.12. Czynności dodatkowe w CCM (rozszerzenie C.6–C.12)</h3>

    <h4>C.6. Zmiana wersji pliku ręcznie wysyłanego</h4>
    <ol class="steps">
      <li>Po otrzymaniu ACK z kodem <code>21 duplicatedVersion</code> ustalić numer ostatnio przyjętej wersji w CCCt Message Viewer.</li>
      <li>Wygenerować nowy plik z numerem wersji o 1 wyższym (zachować pozostałą część maski nazwy bez zmian).</li>
      <li>Wysłać ponownie zgodnie z procedurą C.4 (lub C.5 po CET).</li>
      <li>Po przyjęciu zweryfikować, że na pulpicie widoczny jest nowy numer wersji.</li>
    </ol>

    <h4>C.7. Anulowanie wysyłki w toku</h4>
    <ol class="steps">
      <li>Anulowanie jest możliwe tylko przed zakończeniem etapu „Wysyłanie" w oknie dialogowym (przycisk „Anuluj").</li>
      <li>Po zakończeniu wysyłki anulowanie nie jest możliwe — w razie błędu należy ponowić wysyłkę z poprawionym plikiem (kod 21 duplicatedVersion).</li>
    </ol>

    <h4>C.8. Diagnoza kafelka z „?" (znak zapytania)</h4>
    <ol class="steps">
      <li>Najechać kursorem na kafelek z „?" — w tooltip pojawi się czas oczekiwania na ACK.</li>
      <li>Otworzyć Core CC Tool → Message Viewer → filtr po kodzie przepływu i dacie biznesowej.</li>
      <li>Status <strong>Processed</strong> w CCCt mimo „?" w CCM → odświeżyć pulpit (Ctrl+F5) i zaczekać do najbliższego cyklu odświeżenia.</li>
      <li>Status <strong>Rejected/Error</strong> w CCCt → przejść do scenariusza <a href="#r05">R05</a>.</li>
      <li>Brak wpisu w CCCt po 5 min → eskalacja zgodnie z R07 (brak CCCt) lub R09 (manual upload).</li>
    </ol>

    <h4>C.9. Pobieranie pliku z CCM (re-download)</h4>
    <ol class="steps">
      <li>Kliknąć właściwy kafelek → menu kontekstowe → <strong>„Pobierz"</strong>.</li>
      <li>Plik zapisywany jest do katalogu pobierania przeglądarki pod oryginalną nazwą.</li>
      <li>Opcja dostępna dla plików, które trafiły do CCM przez SFTP/ZP, MinIO lub które zostały ręcznie wysłane przez CCM.</li>
    </ol>

    <h4>C.10. Porównanie wersji (history viewer)</h4>
    <ol class="steps">
      <li>Kliknąć kafelek → menu kontekstowe → <strong>„Historia wersji"</strong>.</li>
      <li>Widok pokazuje listę wszystkich wersji pliku w danej dobie biznesowej z numerem, timestampem dostarczenia i statusem.</li>
      <li>Dla IVA BACKUP — kluczowe narzędzie weryfikacji, czy publikowana wersja odpowiada wersji wyniku z Perun4V (R02/R12).</li>
    </ol>

    <h4>C.11. Filtrowanie i wyszukiwanie</h4>
    <ol class="steps">
      <li>Ctrl+F w pulpicie aktywuje pasek wyszukiwania po nazwie przepływu i kodzie FIDx.</li>
      <li>Filtr „Tylko niezielone" (ikona w nagłówku) ukrywa wiersze ze statusami OK — przydatne w trybie dyżurnym dla szybkiej oceny problemów.</li>
      <li>Filtr „Tylko bieżąca iteracja" — automatycznie zwęża widok do iteracji wynikającej z aktualnego czasu.</li>
    </ol>

    <h4>C.12. Praca równoległa wielu dyspozytorów</h4>
    <div class="box box-warn">
      <strong>⚠ Tryb wykonania</strong>
      Pulpit CCM nie blokuje równoległej pracy dwóch dyspozytorów na tej samej dobie biznesowej. Przed każdą ręczną wysyłką (C.4/C.5) potwierdzić telefonicznie z drugim stanowiskiem, że czynność jest nieduplikowana. Każdą wysyłkę odnotować w dokumentacji dyżurowej z czasem, identyfikatorem operatora i uzasadnieniem.
    </div>

    <h3 id="s7m13" class="anchor-target">7M.13. Macierz uprawnień stanowiskowych</h3>
    <div class="tbl-wrap"><table>
      <thead><tr><th>Rola</th><th>Odczyt pulpitu</th><th>Wysyłka przed CET (C.4)</th><th>Wysyłka po CET (C.5)</th><th>Manual upload CCCt</th><th>Konfiguracja widoku</th></tr></thead>
      <tbody>
        <tr><td>Dyspozytor IDCC (operacyjny)</td><td>✅</td><td>✅</td><td>✅ (z uzasadnieniem)</td><td>✅ (R05/R09)</td><td>✅ (własny widok)</td></tr>
        <tr><td>Dyspozytor wspierający</td><td>✅</td><td>✅ (tylko backup)</td><td>❌</td><td>✅ (po uzgodnieniu)</td><td>✅ (własny widok)</td></tr>
        <tr><td>Kierownik zmiany</td><td>✅</td><td>❌ (nadzór)</td><td>❌ (autoryzacja)</td><td>❌</td><td>✅ (globalna)</td></tr>
        <tr><td>Inżynier CCC / PSE Innowacje</td><td>✅</td><td>❌</td><td>❌</td><td>✅ (interwencja)</td><td>✅ (administracyjna)</td></tr>
      </tbody>
    </table></div>

    <h3 id="s7m14" class="anchor-target">7M.14. Sytuacje awaryjne pulpitu CCM</h3>
    <div class="tbl-wrap"><table>
      <thead><tr><th>Objaw</th><th>Pierwsza diagnoza</th><th>Działanie</th></tr></thead>
      <tbody>
        <tr><td>Pulpit nie ładuje się (biały ekran)</td><td>Ctrl+F5; sprawdzić sieć; sprawdzić, czy https://ccm.spsm.pse.pl/ pinguje</td><td>Eskalacja do PSE Innowacje; w międzyczasie monitoring przez MinIO + Core CC Tool (R06).</td></tr>
        <tr><td>Pulpit ładuje się, ale brak danych dla doby</td><td>Sprawdzić Business Day; sprawdzić tryb odświeżania; sprawdzić, czy inne doby też są puste</td><td>Jeżeli inne doby też puste — eskalacja do PSE Innowacje. Jeżeli tylko bieżąca — zweryfikować w CCCt, czy proces wystartował.</td></tr>
        <tr><td>Kafelki nie aktualizują się mimo trybu ultraszybkiego</td><td>Sprawdzić licznik odświeżenia; sprawdzić tryb kłódki (włączony blokuje akcje, nie aktualizacje)</td><td>Wymusić odświeżenie Ctrl+F5; jeżeli nadal brak — restart przeglądarki; w razie powtarzania — eskalacja PSE Innowacje.</td></tr>
        <tr><td>Kafelek pulsuje, lecz menu kontekstowe nie otwiera się</td><td>Sprawdzić ikonę kłódki (blokada widoku); zalogować się ponownie</td><td>Wyłączyć blokadę; w razie utrzymywania się problemu — restart sesji.</td></tr>
        <tr><td>Komunikat „Sesja wygasła"</td><td>—</td><td>Zalogować ponownie; sprawdzić, czy w międzyczasie nie była wymagana ręczna interwencja (przejrzeć log zdarzeń po zalogowaniu).</td></tr>
      </tbody>
    </table></div>
'''

# ── 3.5. Sekcja 7.8 — realne screenshoty CCM (zastępuje generowane tabele) ─
CCM_DASHBOARD = '''
    <h3 id="s7dash" class="anchor-target">7.8. Wizualizacja pulpitu dyspozytorskiego CCM (rzeczywiste zrzuty ekranu)</h3>

    <div class="box box-info">
      <strong>ℹ Cel sekcji</strong>
      Sekcja prezentuje rzeczywiste zrzuty pulpitu dyspozytorskiego CCM (środowisko produkcyjne) wraz z opisem widocznych elementów i ich znaczenia. Wszystkie zrzuty pochodzą z aplikacji <code>https://ccm.spsm.pse.pl/</code> z pulpitów „IDCC FB Pulpit dyspozytorski" i „IDCC (do 3C) z ATC i SFTP Pulpit dyspozytorski".
    </div>

    <h4>7.8.1. Lista pulpitów — punkt wejścia</h4>
    <p>Po zalogowaniu do CCM dyspozytor wybiera właściwy pulpit z listy. Dla procesu IDCC dostępne są dwa pulpity, w zależności od zakresu monitorowania (FB / FB+ATC+SFTP).</p>
    <figure class="ccm-fig">
      <img src="img_ccm/ccm-001.png" alt="Lista pulpitów CCM" />
      <figcaption>Ekran „Lista pulpitów" — wybór pulpitu odbywa się ikoną oka po lewej stronie nazwy. „DACC FB Pulpit dyspozytorski" obowiązuje od 2024-11-11, „IDCC (do 3C) z ATC i SFTP Pulpit dyspozytorski" od 2025-06-11.</figcaption>
    </figure>

    <h4>7.8.2. Wybór doby biznesowej (Business Day)</h4>
    <p>Selektor Business Day znajduje się w górnym lewym rogu pulpitu. Datę można zmienić strzałkami (◀ ▶) lub kalendarzem. CCM automatycznie przełącza dobę monitorowaną na D w godz. 00:01–10:15 i D+1 w godz. 10:16–00:00.</p>
    <figure class="ccm-fig">
      <img src="img_ccm/ccm-005.png" alt="Selektor Business Day" />
      <figcaption>Selektor Business Day z nawigacją (zaznaczone na czerwono strzałki nawigacyjne).</figcaption>
    </figure>

    <h4>7.8.3. Pełen widok pulpitu IDCC z trybami odświeżania</h4>
    <p>Pulpit „IDCC (do 3C) z ATC i SFTP" prezentuje wszystkie iteracje (IDCC(a), IDCC(b), IDA3, IDCC(c), IDCC(d)) zorganizowane w sekcjach. Widoczne są kolumny statusów: <strong>CCCtool wysł.</strong>, <strong>CCCtool zwalid.</strong>, <strong>MinIO</strong>, <strong>sFTP/ZP wysł.</strong>. W prawym górnym rogu — trzy tryby odświeżania (10 s / 30 s / 5 min), licznik najbliższego odświeżenia, „Legenda", „Pomoc", kłódka i „Test powiadomień".</p>
    <figure class="ccm-fig">
      <img src="img_ccm/ccm-097.png" alt="Pełen widok pulpitu IDCC" />
      <figcaption>Pulpit „IDCC (do 3C) z ATC i SFTP" — wszystkie iteracje z kafelkami statusów. Iteracje rozwijają się symbolem ▼.</figcaption>
    </figure>

    <figure class="ccm-fig">
      <img src="img_ccm/ccm-082.png" alt="Tryby odświeżania i Test powiadomień" />
      <figcaption>Górna belka pulpitu — wybór trybu odświeżania, „Test powiadomień" (potwierdzenie Chrome o poprawnej konfiguracji systemu) oraz ikony pomocnicze (Legenda, Pomoc, Kłódka, Koło zębate).</figcaption>
    </figure>

    <h4>7.8.4. Legenda kolorów statusów (zrzut z pulpitu)</h4>
    <p>Legenda dostępna jest przyciskiem „Legenda" w górnym prawym rogu pulpitu. Pełen opis 14 kategorii kolorystycznych znajduje się także w sekcji <a href="#s7m2">7M.2</a>.</p>
    <figure class="ccm-fig">
      <img src="img_ccm/ccm-017.png" alt="Legenda kolorów CCM" />
      <figcaption>Pełna legenda kolorów dla wszystkich możliwych statusów kafelków (źródło: pulpit CCM produkcja).</figcaption>
    </figure>

    <h4>7.8.5. Rozwinięty agregat składowych paczki FBA</h4>
    <p>Po kliknięciu strzałki ▼ przy agregacie (np. <code>PERUN_IDx</code>) pulpit pokazuje wszystkie składowe paczki w wierszach podrzędnych. Każda składowa ma własny status w kolumnie <strong>MinIO</strong>. Status agregatu zależy od kompletności składowych.</p>
    <figure class="ccm-fig">
      <img src="img_ccm/ccm-135.png" alt="Rozwinięty agregat paczki FBA IDCC(b)" />
      <figcaption>Rozwinięty agregat IDCC(b) — Individual Validation. Widoczne składowe: FID2-620 (CGM), FID2-632 (RefProg), FID2-645 (Initial FB Domain), FID2-657 (Real GLSK), FID2-659 (CGM DQ Check), FID2-665 (Filtered CB), FID2-701 (Vertices), XID_RA (Remedial Actions). TET 20:40, CET 21:05. Większość zielonych (W) — ręczne dostarczenie, jedna składowa w trybie oczekiwania.</figcaption>
    </figure>

    <h4>7.8.6. Menu kontekstowe kafelka</h4>
    <p>Kliknięcie kafelka statusu otwiera menu kontekstowe z opcjami: <strong>„informacje"</strong> (rozmiar pliku, lokalizacja, data dostarczenia), <strong>„pobierz"</strong> (re-download), <strong>„wyślij" / „wyślij po CET"</strong> (ręczne dostarczenie). Dostępność opcji zależy od statusu kafelka, kolumny i upływu CET.</p>
    <figure class="ccm-fig">
      <img src="img_ccm/ccm-138.png" alt="Menu kontekstowe kafelka CCM" />
      <figcaption>Menu kontekstowe kafelka (FID2-665, kolumna MinIO) z opcjami „informacje", „pobierz", „wyślij po CET" — przykład w IDCC(b), Individual Validation.</figcaption>
    </figure>

    <h4>7.8.7. Okno wysyłki pliku (drag &amp; drop)</h4>
    <p>Wybranie opcji „wyślij" otwiera okno modalne „Wyślij [kod pliku]" z polem drag&amp;drop. Plik można przeciągnąć lub wskazać z dysku. Po załadowaniu CCM waliduje nazwę pliku (zgodność z maską FIDx, doba biznesowa, wersja). W razie błędu walidacji wyświetlany jest komunikat z kodem (np. „Niepoprawna nazwa pliku — kod błędu: 40").</p>
    <figure class="ccm-fig">
      <img src="img_ccm/ccm-111.png" alt="Okno wysyłki pliku FID2-657" />
      <figcaption>Okno „Wyślij FID2-657" — drag&amp;drop z listą plików dostępnych na dysku (po prawej Microsoft Edge eksplorator z plikami 20240515-FID2-…).</figcaption>
    </figure>

    <figure class="ccm-fig">
      <img src="img_ccm/ccm-124.png" alt="Okno wysyłki pliku z błędem walidacji nazwy" />
      <figcaption>Okno „Wyślij FID1-250 po CET" z błędem walidacji: „Niepoprawna nazwa pliku — kod błędu: 40". Data biznesowa w nazwie pliku (20240501) nie odpowiada dobie monitorowanej (2025-07-01).</figcaption>
    </figure>

    <h4>7.8.8. Monitoring IVA i IVA Backup</h4>
    <p>Wiersze IVA (FIDx-710) i IVA Backup widoczne są w sekcji Individual Validation. Wersja widoczna w kolumnie „Wersja". Próba ręcznej wysyłki IVA Backup jest blokowana — CCM wyświetla komunikat „Wysyłka backup nie jest możliwa — wersja dokumentu musi być wyższa od docelowej wersji".</p>
    <figure class="ccm-fig">
      <img src="img_ccm/ccm-149.png" alt="Monitoring IVA i IVA Backup IDCC(b)" />
      <figcaption>Pulpit IDCC(b) — wiersze IVA i IVA Backup (FID2-710, wersja 2). W lewym dolnym rogu komunikat blokujący próbę ręcznej wysyłki backup. Wiersz AC (FID2-831) z czerwonym „?" — brak ACK po wysyłce.</figcaption>
    </figure>

    <div class="box box-warn">
      <strong>⚠ Uwagi do zrzutów ekranu</strong>
      <ul>
        <li>Wszystkie zrzuty pochodzą z istniejących sesji produkcyjnych (lata 2024–2026). Daty biznesowe na zrzutach mają wyłącznie charakter ilustracyjny.</li>
        <li>Daty w nazwach plików (np. <code>20240515-FID2-…</code>) służą jako przykład maski nazwy pliku (patrz sekcja 10.2).</li>
        <li>Widoczne komunikaty błędów (kod 40 itp.) oraz pulsacje kafelków są zgodne z opisem w sekcji <a href="#s7m2">7M.2</a> i kodami ACK w <a href="#s7m8">7M.8</a>.</li>
      </ul>
    </div>
'''

# ── 4. Build HTML ──────────────────────────────────────────────────────────
src = SRC.read_text(encoding='utf-8')

# 4a. Usuń wbudowany editor-bar v9 z body (występuje w obu wersjach — naszą edycję dodamy później)
def strip_div_block(src, start_re):
    m = re.search(start_re, src)
    if not m: return src
    i = m.start()
    depth = 0
    j = i
    for mm in re.finditer(r'<(/?)(div|nav)[^>]*>', src[i:], re.I):
        if mm.group(1) == '': depth += 1
        else:
            depth -= 1
            if depth == 0:
                j = i + mm.end()
                break
    return src[:i] + src[j:]

src = strip_div_block(src, r'<div class="editor-bar"')
src = strip_div_block(src, r'<div id="search-panel"')
# usuń też panel search/replace edytora (jeżeli osobno)
src = re.sub(r'<div class="sr-[^"]*"[^>]*>.*?</div>', '', src, flags=re.S)
src = re.sub(r'<div class="find-replace[^"]*"[^>]*>.*?</div>', '', src, flags=re.S)
# usuń skrypty edytora v9 (wszystkie <script> z body — i tak dodamy własne dla edytora)
src = re.sub(r'<script\b[^>]*>.*?</script>', '', src, flags=re.S|re.I)

# 4a2. Zastąp legendę 7.2 (CSS-kropki) realnymi ikonami z pulpitu
LEGEND_72 = '''<div class="status-grid">
      <div class="status-item">''' + color_dot('zielony', 24) + '''<span><strong>Zielony</strong> — dokument został prawidłowo obsłużony</span></div>
      <div class="status-item">''' + color_dot("zielony 'R'", 24) + '''<span><strong>Zielony R</strong> — dokument wysłany ręcznie</span></div>
      <div class="status-item">''' + color_dot("ciemnozielony 'W'", 24) + '''<span><strong>Zielony W</strong> — dokument/składowa agregatu wysłany z CCM przed CET</span></div>
      <div class="status-item">''' + color_dot("ciemnozielony 'W po CET'", 24) + '''<span><strong>Ciemnozielony W</strong> — dokument/składowa agregatu wysłany z CCM po CET</span></div>
      <div class="status-item">''' + color_dot('zielony z pulsacją', 24) + '''<span><strong>Zielony z pulsującą obwódką</strong> — wymagana akcja Użytkownika (np. wysyłka Backup do CCTool)</span></div>
      <div class="status-item">''' + color_dot('pomarańczowy (TET)', 24) + '''<span><strong>Pomarańczowy</strong> — zbliża się Target End Time; brak dokumentu</span></div>
      <div class="status-item">''' + color_dot('czerwony (CET)', 24) + '''<span><strong>Czerwony</strong> — zbliża się Critical End Time; brak dokumentu lub błąd</span></div>
      <div class="status-item">''' + color_dot("czerwony z '?'", 24) + '''<span><strong>Czerwony ?</strong> — proces się nie skończył lub brak dokumentu ACK</span></div>
      <div class="status-item">''' + color_dot("czerwony z '!'", 24) + '''<span><strong>Czerwony !</strong> — niepoprawny rozmiar pliku</span></div>
      <div class="status-item">''' + color_dot('fioletowy', 24) + '''<span><strong>Fioletowy</strong> — dostarczony po CET</span></div>
      <div class="status-item">''' + color_dot('czarny', 24) + '''<span><strong>Czarny</strong> — dokument niedostarczony</span></div>
      <div class="status-item">''' + color_dot('szary ?', 24) + '''<span><strong>Szary ?</strong> — proces w trakcie wykonywania</span></div>
      <div class="status-item">''' + color_dot('szary', 24) + '''<span><strong>Jasnoszary</strong> — dokument oczekujący na obsługę</span></div>
    </div>'''
m_lg = re.search(r'<div class="status-grid">.*?\n    </div>', src, re.S)
if m_lg:
    src = src[:m_lg.start()] + LEGEND_72 + src[m_lg.end():]

# 4a3. Aktualizacja metryki dokumentu (wersja/data)
src = src.replace('<dd>5.1</dd>', '<dd>5.2</dd>')
src = src.replace('21 kwietnia 2026', '10 czerwca 2026')
src = src.replace('v5.1 · 21.04.2026', 'v5.2 · 10.06.2026')
src = src.replace('v5.1', 'v5.2')

# 4b. Zastąp sekcję 7.8 (wizualizacja dashboard) realnymi screenshotami
# Stara sekcja: od `<h3 id="s7dash"` do bezpośrednio przed `<!-- ═══ SEKCJA 8`
m_start = re.search(r'<h3 id="s7dash"[^>]*>', src)
m_end   = re.search(r'<!-- ═══ SEKCJA 8', src)
if m_start and m_end:
    src = src[:m_start.start()] + CCM_DASHBOARD + '\n\n    ' + src[m_end.start():]

# Dodatkowe style dla nowej sekcji
extra_css = '''
    /* ── ROZSZERZENIE 7M ── */
    .sdot { vertical-align:middle; margin-right:5px; border-radius:4px; }
    .cdot-spin { display:inline-block; width:16px; height:16px; border-radius:50%; border:2px solid #BFCFDF; border-top-color:#0070C0; vertical-align:middle; margin-right:5px; animation:spin 1s linear infinite; }
    @keyframes spin { to { transform:rotate(360deg); } }
    .tile-state { display:flex; align-items:center; gap:4px; margin:2px 0; white-space:nowrap; font-size:11px; }
    .subhead { font-size:14px; color:var(--pse-blue); margin:20px 0 8px; padding:4px 0; border-bottom:1px dashed var(--border); }
    table.tile-cat { table-layout:fixed; width:100%; }
    table.tile-cat th, table.tile-cat td { font-size:11.5px; padding:4px 6px; overflow-wrap:break-word; }
    table.state-cat { table-layout:fixed; width:100%; }
    table.state-cat th, table.state-cat td { font-size:11.5px; padding:5px 7px; vertical-align:top; overflow-wrap:break-word; }
    .kat-badge { display:inline-block; min-width:22px; text-align:center; color:#fff; font-size:10px; font-weight:700; padding:2px 5px; border-radius:3px; margin-right:4px; }
    .fid-card { border:1px solid var(--border); border-radius:4px; margin:8px 0; background:#fff; page-break-inside:avoid; }
    .fid-head { background:var(--pse-blue-light); padding:6px 10px; border-bottom:1px solid var(--border); display:flex; flex-wrap:wrap; gap:10px; align-items:center; font-size:12px; }
    .fid-code { background:var(--pse-blue); color:#fff; font-family:monospace; font-weight:700; padding:2px 8px; border-radius:3px; font-size:11px; }
    .fid-name { font-weight:600; color:var(--pse-blue); }
    .fid-meta { margin-left:auto; font-size:11px; color:var(--pse-gray); }
    .fid-body { padding:8px 12px; font-size:12px; }
    .fid-body p { margin:3px 0; }
    .fid-body .uw { color:var(--pse-orange); }
    @keyframes pulse { 0%,100% { box-shadow: 0 0 0 3px rgba(192,0,0,.4);} 50% { box-shadow: 0 0 0 6px rgba(192,0,0,.1);} }

    /* ── ZRZUTY EKRANU CCM ── */
    .ccm-fig { margin:18px 0; padding:10px; border:1px solid var(--border); border-radius:5px; background:#fafcfe; text-align:center; page-break-inside:avoid; }
    .ccm-fig img { max-width:100%; height:auto; border:1px solid #D0DAE5; border-radius:3px; box-shadow:0 1px 4px rgba(0,0,0,.08); }
    .ccm-fig figcaption { margin-top:8px; font-size:11.5px; color:var(--pse-gray); font-style:italic; line-height:1.45; text-align:left; padding:0 6px; }
'''

# wstaw extra_css przed </style>
src = src.replace('</style>', extra_css + '\n  </style>', 1)

# Wstaw nową sekcję 7M PRZED sekcją 8
ins_point = '<!-- ═══ SEKCJA 8 — MinIO/Perun/CCCt'
if ins_point not in src:
    raise RuntimeError('Nie znaleziono punktu wstawienia')
src = src.replace(ins_point, CCM_EXT + '\n\n    ' + ins_point)

# Zaktualizuj sidebar — dodaj linki do nowej sekcji 7M
# Znajdź sekcję 7 w sidebar i wstaw 7M poniżej
sidebar_addition = '''        <a href="#s7m" class="nav-item">7M. Katalog CCM (rozsz.)</a>
        <a href="#s7m2" class="nav-sub">7M.2 Legenda 14 statusów</a>
        <a href="#s7m4" class="nav-sub">7M.4 Katalog kafelków</a>
        <a href="#s7m5" class="nav-sub">7M.5 Katalog plików FIDx</a>
        <a href="#s7m6" class="nav-sub">7M.6 Katalog stanów</a>
        <a href="#s7m7" class="nav-sub">7M.7 Scenariusze S/R</a>
        <a href="#s7m8" class="nav-sub">7M.8 Kody ACK</a>
        <a href="#s7m12" class="nav-sub">7M.12 Czynności C.6–C.12</a>
        <a href="#s7m14" class="nav-sub">7M.14 Sytuacje awaryjne CCM</a>
'''
# Wstaw przed pierwszym wystąpieniem '#s8' w sidebar nav
m = re.search(r'(<a href="#s8"[^>]*>)', src)
if m:
    src = src[:m.start()] + sidebar_addition + '        ' + src[m.start():]

# Zaktualizuj numer wersji w cover (v5.1 → v5.2)
src = src.replace('v5.1', 'v5.2', 1)
src = src.replace('PROCEDURA_IDCC_TSO_v5_1', 'PROCEDURA_IDCC_TSO_v5_2')

# Zaktualizuj tytuł
src = src.replace('<title>FBA_TSO_IDCC – Procedura procesu IDCC | PSE S.A.</title>',
                  '<title>FBA_TSO_IDCC v5.2 – Procedura procesu IDCC + pełny katalog CCM | PSE S.A.</title>')

# ── 5. Zapisz wersję bez edytora ──────────────────────────────────────────
out_plain = ROOT / 'PROCEDURA_IDCC_TSO_v5_2.html'
out_plain.write_text(src, encoding='utf-8')
print(f'Zapisano: {out_plain}  ({len(src):,} bajtów, {src.count(chr(10)):,} linii)')

# ── 6. Wersja z edytorem ──────────────────────────────────────────────────
editor_css = '''
    /* ── EDITOR MODE ── */
    .editor-toolbar { position:fixed; top:0; left:260px; right:0; height:42px; background:#1a1a1a; color:#fff; display:flex; align-items:center; padding:0 16px; gap:8px; z-index:200; box-shadow:0 2px 6px rgba(0,0,0,.3); font-size:12px; }
    .editor-toolbar button { background:#0070C0; color:#fff; border:0; padding:6px 12px; border-radius:3px; cursor:pointer; font-size:11.5px; font-weight:600; }
    .editor-toolbar button:hover { background:#005AA9; }
    .editor-toolbar button.warn { background:#D46B08; }
    .editor-toolbar .status { margin-left:auto; opacity:.7; font-size:11px; }
    body.editor-on .main { padding-top:42px; }
    body.editor-on [contenteditable="true"]:focus { outline:2px solid #0070C0; outline-offset:2px; background:#FFFEF0; }
    body.editor-on .editable { border-left:2px dotted #BFCFDF; }
'''

editor_toolbar = '''
  <div class="editor-toolbar">
    <strong>📝 Tryb edycji</strong>
    <button onclick="toggleEdit()">Włącz/Wyłącz edycję</button>
    <button onclick="saveLocal()">💾 Zapisz lokalnie</button>
    <button onclick="exportHTML()" class="warn">📤 Eksportuj HTML</button>
    <button onclick="printDoc()">🖨 Drukuj/PDF</button>
    <span class="status" id="ed-status">Edycja: OFF · Ostatni zapis: —</span>
  </div>
'''

editor_script = '''
<script>
let editOn = false;
function toggleEdit() {
  editOn = !editOn;
  document.body.classList.toggle('editor-on', editOn);
  document.querySelectorAll('.main h1, .main h2, .main h3, .main h4, .main h5, .main p, .main li, .main td, .main dd').forEach(el => {
    el.setAttribute('contenteditable', editOn ? 'true' : 'false');
    if (editOn) el.classList.add('editable'); else el.classList.remove('editable');
  });
  document.getElementById('ed-status').textContent = 'Edycja: ' + (editOn ? 'ON' : 'OFF') + ' · Ostatni zapis: ' + (localStorage.getItem('idcc_v52_ts') || '—');
}
function saveLocal() {
  const html = document.documentElement.outerHTML;
  try {
    localStorage.setItem('idcc_v52_doc', html);
    const ts = new Date().toLocaleString('pl-PL');
    localStorage.setItem('idcc_v52_ts', ts);
    document.getElementById('ed-status').textContent = 'Edycja: ' + (editOn ? 'ON' : 'OFF') + ' · Ostatni zapis: ' + ts;
    alert('Zapisano lokalnie w przeglądarce (localStorage).');
  } catch (e) { alert('Błąd zapisu: ' + e.message); }
}
function exportHTML() {
  const html = '<!DOCTYPE html>\\n' + document.documentElement.outerHTML;
  const blob = new Blob([html], {type:'text/html;charset=utf-8'});
  const url = URL.createObjectURL(blob);
  const a = document.createElement('a');
  a.href = url; a.download = 'PROCEDURA_IDCC_TSO_v5_2_edited.html';
  document.body.appendChild(a); a.click(); a.remove();
  URL.revokeObjectURL(url);
}
function printDoc() { window.print(); }
// Wczytanie ostatniej wersji
window.addEventListener('DOMContentLoaded', () => {
  const saved = localStorage.getItem('idcc_v52_doc');
  if (saved && confirm('Wykryto wcześniej zapisaną wersję dokumentu. Czy wczytać?')) {
    document.open(); document.write(saved); document.close();
  }
});
</script>
'''

src_ed = src.replace('</body>', editor_script + '\n</body>')
# Wstaw toolbar po <body
src_ed = re.sub(r'(<body[^>]*>)', r'\1\n' + editor_toolbar, src_ed, count=1)
# Dodaj CSS edytora
src_ed = src_ed.replace('</style>', editor_css + '\n  </style>', 1)
# Zaktualizuj tytuł
src_ed = src_ed.replace('v5.2 – Procedura procesu', 'v5.2 [EDYTOR] – Procedura procesu', 1)

out_ed = ROOT / 'PROCEDURA_IDCC_TSO_v5_2_EDYTOR.html'
out_ed.write_text(src_ed, encoding='utf-8')
print(f'Zapisano: {out_ed}  ({len(src_ed):,} bajtów)')
