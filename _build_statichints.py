#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generuje staticHints (TS) dla WSZYSTKICH plików — w formacie pliku-wzoru
   CCM/staticHints-przyklad-dla-FID1-831.txt.

Źródła:
  Stany_kafelkow_per_plik_IDCC.xlsx   — stany per kod (markery + fill = kafelek),
                                         pełne ramki (sytuacja, procedury, opis…)
  CCM/staticHints-przyklad-dla-FID1-831.txt — wzór formatu + enumy AC (do walidacji
                                         i tabeli sygnatur wierszy dla profilu 2CCT)

Mapowanie kafelek → DashboardItemStatus:
  • 8 kombinacji (marker,fill) jednoznacznych — wprost.
  • wieloznaczne (czerwony/fioletowy/czarny bez markera) → tabela sygnatur całego
    wiersza wyprowadzona z AC (dokładna dla 2CCT/3COL o tych samych sygnaturach),
    z fallbackiem per-cela.
  • pomarańczowy (zbliża się TET) i szary‑? — brak w pliku-wzorze → oznaczone /*?*/
    do potwierdzenia z definicją enuma w programie.

Wyjście: CCM/staticHints-IDCC-all.txt
"""
import re
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
CCM  = ROOT / 'CCM'
XLSX = ROOT / 'Stany_kafelkow_per_plik_IDCC.xlsx'
SAMPLE = CCM / 'staticHints-przyklad-dla-FID1-831.txt'

SHEETS = ['AC 831 (2CCT)','GLSK CBCORA 607-617 (2CCT)','ATC 928 (zwrotny)',
          'IVA 710 (3COL)','Paczki PERUN (2MS)','Skladowe MinIO (MIN)',
          'sFTP-ZP NTC MD250 IGM (SFTP)']
COLNAME = {1:'CCtool wysł.', 2:'CCtool zwalid.', 3:'MinIO', 4:'sFTP/ZP wysł.'}

# Kombinacje niemożliwe w produkcji (feedback z 14.07):
# GLSK: nie występują SUCCESS+FAILURE / SUCCESS+ACKNOTRECEIVED / SUCCESS+EXCEEDED
# RefProg: nie występuje W (SENTMANUALLY) — mamy szary+?
# sFTP: nie występuje W (SENTMANUALLY)
IMPOSSIBLE = [
    # (file_codes, status_combo) — file_codes = zbiór kodów plików, gdzie combo jest niemożliwe
    (frozenset(['FID2-607','FID3-607','FID3C-607']),
     frozenset([('CCtool wysł.','SUCCESS'),('CCtool zwalid.','FAILURE')])),
    (frozenset(['FID2-607','FID3-607','FID3C-607']),
     frozenset([('CCtool wysł.','SUCCESS'),('CCtool zwalid.','ACKNOTRECEIVED')])),
    (frozenset(['FID2-607','FID3-607','FID3C-607']),
     frozenset([('CCtool wysł.','SUCCESS'),('CCtool zwalid.','EXCEEDED')])),
    (frozenset(['FID2-632','FID3-632','FID3C-632']),
     frozenset([('MinIO','SENTMANUALLY')])),
    (frozenset(['FID2-632','FID3-632','FID3C-632']),
     frozenset([('MinIO','SENTMANUALLYAFTERCET')])),
    (frozenset(['FID1-921','FID2-921','FID3-921','FID3C-921']),
     frozenset([('sFTP/ZP wysł.','SENTMANUALLY')])),
    (frozenset(['FID1-921','FID2-921','FID3-921','FID3C-921']),
     frozenset([('sFTP/ZP wysł.','SENTMANUALLYAFTERCET')])),
]

# ── per-cela: (marker, fill) → DashboardItemStatus ─────────────────────────
# Mapowanie 1:1 wg OFICJALNEJ legendy enumów programu (kafelek → status):
CELL_ENUM = {
    ('·','00FFFFFF'):'UNKNOWN',            # 1  białe/puste
    ('·','00BFBFBF'):'UNKNOWN',            # 1  szary
    ('·','00C00000'):'FAILURE',            # 2  czerwony
    ('·','0000B050'):'SUCCESS',            # 3  zielony
    ('R','0000B050'):'FORCEDSUCCESS',      # 4  zielony R
    ('W','002E7D32'):'SENTMANUALLY',       # 5  ciemnozielony W
    ('W↑CET','001B5E20'):'SENTMANUALLYAFTERCET',  # 6  W po CET
    ('?','00C00000'):'ACKNOTRECEIVED',     # 7  czerwony ?
    ('·','00ED7D31'):'WARNING',            # 8  pomarańczowy
    ('·','00000000'):'EXCEEDED',           # 9  czarny
    ('!','00C00000'):'FILESIZEBAD',        # 10 czerwony !
    ('?','00BFBFBF'):'PROCESSRUNNING',     # 11 szary ?
    ('·','007030A0'):'CONN_AFTER_CET',     # 12 fioletowy
}
FLAG_COMBO = set()   # brak — mapowanie pełne wg legendy

def cellkey(c):
    if c is None: return None
    if str(c.value) == '—': return None
    if not (c.fill and c.fill.patternType): return None   # brak wypełnienia = kolumna nieaktywna
    fill = c.fill.fgColor.rgb
    if fill is None: return None
    m = '·' if c.value in (None,'None','') else str(c.value)
    return (m, fill)

# ── tabela sygnatur wierszy z AC (exact dla 2CCT) ──────────────────────────
def load_ac_signatures():
    txt = SAMPLE.read_text(encoding='utf-8')
    blocks = re.split(r'\n\s*\{\s*\n\s*status:', txt)[1:]
    enums = [re.findall(r"DashboardItemStatus\.([A-Z_]+)", b) for b in blocks]
    wb = load_workbook(XLSX)
    rows = [r for r in wb['AC 831 (2CCT)'].iter_rows(min_row=2) if r[0].value=='FID1-831']
    sig = {}
    for i, r in enumerate(rows):
        key = tuple(cellkey(r[ci]) for ci in (1,2,3,4) if cellkey(r[ci]))
        if i < len(enums):
            sig[key] = tuple(enums[i])
    return sig

AC_SIG = load_ac_signatures()

def row_enums(cells):
    """cells: lista (colName, cellkey) tylko aktywnych kolumn → lista enumów.
       Spójnie: enum per-cela dla każdej aktywnej kolumny."""
    return [CELL_ENUM.get(ck, 'UNKNOWN') for _, ck in cells]

# ── parser pełnej ramki ────────────────────────────────────────────────────
FRAME_SECTIONS = ['Stan kafelków','Opis stanu kafelków','Procedury awaryjne',
                  'Kafelek MinIO / sFTP','Opis pliku','Ścieżka pliku',
                  'Kontakt w przypadku awarii']
def parse_frame(text):
    if not text: return {}
    text = text.replace('\r\n','\n')
    body = '\n'.join(text.split('\n')[1:])
    idx = []
    for sec in FRAME_SECTIONS:
        for mm in re.finditer(r'(?m)^'+re.escape(sec)+r'\s*$', body):
            idx.append((mm.start(), mm.end(), sec))
    idx.sort()
    rec = {}
    for i,(s,e,sec) in enumerate(idx):
        nxt = idx[i+1][0] if i+1 < len(idx) else len(body)
        rec[sec] = body[e:nxt].strip()
    return rec

def inline(s):
    """Zwija tekst do jednej linii i usuwa konkretne godziny przy TET/CET.
       Czasy różnią się per plik i są widoczne na pulpicie — zostawiamy samo „TET"/„CET"."""
    s = re.sub(r'[ \t]*\n[ \t]*', ' ', str(s or '')).strip()
    s = re.sub(r'\b(TET|CET)\s*\d{1,2}[:.]\d{2}', r'\1', s)   # „czekaj do TET 04:15" → „czekaj do TET"
    s = re.sub(r'\s+([.,;])', r'\1', s)
    return re.sub(r'\s+', ' ', s)

def q(s):
    """Escape do literału w pojedynczych cudzysłowach (TS)."""
    return str(s or '').replace('\\', '\\\\').replace("'", "\\'")

# akcja ratunkowa = to faktyczna procedura (zostaje inline); jej brak = HTML <strong>
ACTION_RE = re.compile(
    r'(→|\bgdy\b|sprawdź|pobierz|wyślij|manual upload|zgłoś|weryfik|zweryfik|'
    r'ustaw|podnieś|telefon|diagnost|wgraj|wygeneruj|reaguj|ponów|'
    r'oznacz|odczytaj|przejdź|otwórz|\bstop\b)', re.I)

# narzędzia/węzły → czytelne hiperłącza do inwentarza (najdłuższe dopasowania pierwsze)
TOOL_LINKS = [
    ('Core CC Tool (Manual Upload)', 'Inwentarz_IDCC.html#P01'),
    ('Manual Upload',                'Inwentarz_IDCC.html#P01'),
    ('Message Viewer',               'Inwentarz_IDCC.html#N02'),
    ('Core CC Tool',                 'Inwentarz_IDCC.html#N02'),
    ('Connector 2.0',                'Inwentarz_IDCC.html#N05'),
    ('Perun4V',                      'Inwentarz_IDCC.html#N03'),
    ('aplikacji ZP',                 'Inwentarz_IDCC.html#N06'),
    ('Kreator',                      'Inwentarz_IDCC.html#N08'),
    ('MinIO',                        'Inwentarz_IDCC.html#N04'),
    ('RCC',                          'Inwentarz_IDCC.html#N13'),
    ('CCM',                          'Inwentarz_IDCC.html#N01'),
]

def linkify(t):
    """Czysty tekst — BEZ linków zewnętrznych i wewnętrznych (wymóg Łukasza).
       Odnośniki do ryzyk zamienione na czytelny tekst bez numerków i nawiasów."""
    # ryzyka → czytelny tekst (bez <a>, bez numerków/nawiasów)
    t = re.sub(r'Szczegóły \(opis \+ screeny\):\s*R\d+(?:\s*·\s*R\d+)*',
               'szczegóły w bazie ryzyk', t)
    t = re.sub(r'\[?\s*R\d+\s*—\s*pełny opis\s*\]?', 'pełny opis w bazie ryzyk', t)
    t = re.sub(r'\bR\d+(?:\s*·\s*R\d+)+', 'powiązane ryzyka w bazie ryzyk', t)
    # odnośniki do innych stanów → czytelnie, bez numerków
    t = re.sub(r'\bStan \d+\s*\(([^)]+)\)', r'stan \1', t)
    t = re.sub(r'(→\s*)Stan \d+\b', r'\1powiązany stan', t)
    t = re.sub(r'\bStan \d+\b', 'powiązany stan', t)
    # znacznik „HTML §x.x" → usuń (zostaw opis)
    t = re.sub(r'HTML §[\d.]+:?\s*', '', t)
    return t

def linkify_path(t):
    """Ścieżka pliku: usuwa kody węzłów [Nxx]/[Pxx] → czysty, czytelny tekst
       (bez linków — CCM nie hostuje Inwentarz_IDCC.html)."""
    t = re.sub(r'\s*\[[NP]\d+\]', '', inline(t))
    return re.sub(r'\s+', ' ', t).strip()

def esc_html(s):
    return s.replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')

def format_proc(text):
    """Tekst → hierarchia (poziom, treść, nagłówek): 0 nagłówek, 1 punkt •, 2 zagnieżdżony 'kod NN'."""
    t = re.sub(r'[ \t]*\n[ \t]*', ' ', str(text or '')).strip()
    t = re.sub(r'\s+', ' ', t)
    t = re.sub(r'\s*(Jeżeli:)', r' ‖\1', t)
    parts = t.split('•')
    out = []
    for seg in re.split(r'‖', parts[0].strip()):
        seg = seg.strip()
        if seg: out.append((0, seg, seg.endswith(':')))
    for it in [p.strip() for p in parts[1:] if p.strip()]:
        sub = re.split(r'‖', it)
        main = sub[0].strip()
        if main: out.append((2 if re.match(r'^kod\s', main, re.I) else 1, main, False))
        for ex in sub[1:]:
            ex = ex.strip()
            if ex: out.append((0, ex, ex.endswith(':')))
    return out

def proc_field(text):
    """emergencyProcedures jako HTML: nagłówki <strong>, punktory <ul><li> (także zagnieżdżone),
       bez linków. Stany pasywne („Brak działań"/„Monitoring"/„…OK") → <strong>…</strong>."""
    t = linkify(inline(text))
    if ('•' not in t) and not ACTION_RE.search(t):
        return f'<strong>{esc_html(t)}</strong>'
    html, depth = [], 0     # depth: 0 brak listy, 1 ul, 2 zagnieżdżony ul
    def close(to):
        nonlocal depth
        while depth >= to and depth > 0:
            html.append('</li></ul>'); depth -= 1
    for lvl, body, hdr in format_proc(t):
        b = esc_html(body)
        if hdr or lvl == 0:
            close(1)
            html.append(f'<strong>{b}</strong>' if hdr else f'<p>{b}</p>')
        elif lvl == 1:
            close(2)
            if depth == 0: html.append('<ul>'); depth = 1
            else: html.append('</li>')
            html.append(f'<li>{b}')
        elif lvl == 2:
            if depth < 1: html.append('<ul>'); depth = 1; html.append('<li>')
            if depth < 2: html.append('<ul>'); depth = 2
            else: html.append('</li>')
            html.append(f'<li>{b}')
    close(1)
    return ''.join(html)

# nazwa przepływu z nagłówka ramki: „KOD — Stan N · KAT — NAZWA"
def frame_name(text):
    head = str(text or '').replace('\r\n','\n').split('\n')[0]
    m = re.match(r'^.*?—\s*Stan\s+\d+\s*·\s*[A-D]\s*—\s*(.*)$', head)
    return m.group(1).strip() if m else ''

def hint_obj(r, code):
    """Buduje jeden obiekt Hint (jako tekst) z wiersza arkusza + jego sygnaturę do dedup."""
    fr = parse_frame(r[5].value)
    active = [(COLNAME[c], cellkey(r[c])) for c in (1,2,3,4) if cellkey(r[c])]
    enums = row_enums(active)

    # sprawdź blacklistę niemożliwych combo
    combo = frozenset((cn, en) for (cn, _), en in zip(active, enums))
    for bad_codes, bad_combo in IMPOSSIBLE:
        if code in bad_codes and combo == bad_combo:
            return None  # pomijamy

    status_lines, flagged = [], 0
    status_lines, flagged = [], 0
    for (cn, ck), en in zip(active, enums):
        flag = '  /*? potwierdź enum*/' if ck in FLAG_COMBO else ''
        if ck in FLAG_COMBO: flagged += 1
        status_lines.append((cn, en, flag))
    sit  = q(inline(fr.get('Opis stanu kafelków','')))
    proc = q(proc_field(fr.get('Procedury awaryjne','')))
    fdesc = q(inline(fr.get('Opis pliku','')))
    fpath = q(linkify_path(fr.get('Ścieżka pliku','')))
    # sygnatura do dedup: statusy + opis stanu + procedura (bez fileDesc, który ma kod FID)
    sig = (tuple((cn, en) for cn, en, _ in status_lines), sit, proc)
    block = ['      {', '        status: [']
    for i,(cn,en,flag) in enumerate(status_lines):
        comma = ',' if (i < len(status_lines)-1 or flag) else ''
        block.append(f"          {{ colName: '{cn}', status: DashboardItemStatus.{en} }}{comma}{flag}")
    block += ['        ],',
              f"        situationDescription: '{sit}',",
              f"        emergencyProcedures: '{proc}',",
              f"        fileDesc: '{fdesc}',",
              f"        filePath: '{fpath}'",
              '      },']
    return '\n'.join(block), sig, flagged

# ── budowa pliku ───────────────────────────────────────────────────────────
def build():
    wb = load_workbook(XLSX)
    # zbierz wiersze per kod, w kolejności wystąpienia
    order, rows_by_code = [], {}
    for shn in SHEETS:
        for r in wb[shn].iter_rows(min_row=2):
            if not r[0].value: continue
            code = str(r[0].value).strip()
            if code not in rows_by_code:
                rows_by_code[code] = []; order.append(code)
            rows_by_code[code].append(r)

    # klucz Record = NAZWA^KOD (np. „CBCORA^FID2-617") — zawsze unikalny per plik
    out = ['  private staticHints: Record<string, Hint[]> = {']
    total_states = flagged_total = 0
    for code in order:
        nm = frame_name(rows_by_code[code][0][5].value) or code
        key = f'{nm}^{code}'
        out.append(f"    '{q(key)}': [")
        seen = set()
        skipped = 0
        for r in rows_by_code[code]:
            result = hint_obj(r, code)
            if result is None:
                skipped += 1
                continue
            block, sig, fl = result
            if sig in seen: continue            # dedup identycznych stanów (np. IDCC(b)+IDA3 po usunięciu godzin)
            seen.add(sig)
            out.append(block)
            total_states += 1; flagged_total += fl
        out.append('    ],')
    out.append('  }')
    text = '\n'.join(out) + '\n'

    dst = CCM / 'staticHints-IDCC-all.txt'
    dst.write_text(text, encoding='utf-8')
    print(f'Zapisano: {dst}')
    print(f'Kluczy (NAZWA^KOD): {len(order)} | stanów łącznie: {total_states} | '
          f'pominięte (niemożliwe): {skipped} | oznaczonych /*?*/: {flagged_total}')
    return dst

# ── walidacja enumów na AC (musi być 100%) ─────────────────────────────────
def validate():
    txt = SAMPLE.read_text(encoding='utf-8')
    blocks = re.split(r'\n\s*\{\s*\n\s*status:', txt)[1:]
    exp = [re.findall(r"DashboardItemStatus\.([A-Z_]+)", b) for b in blocks]
    wb = load_workbook(XLSX)
    rows = [r for r in wb['AC 831 (2CCT)'].iter_rows(min_row=2) if r[0].value=='FID1-831']
    # dokładność per-enum: porównaj enum dla każdej kolumny, którą wzór wymienia
    tot = hit = 0
    arr_ok = 0
    for i, r in enumerate(rows):
        active = [(COLNAME[c], cellkey(r[c])) for c in (1,2,3,4) if cellkey(r[c])]
        got = row_enums(active)
        gmap = {cn: e for (cn, _), e in zip(active, got)}
        # mapa oczekiwana per colName ze wzoru
        b = blocks[i]
        ecols = re.findall(r"colName:\s*'([^']+)',\s*status:\s*DashboardItemStatus\.([A-Z_]+)", b)
        all_match = True
        for cn, en in ecols:
            tot += 1
            if gmap.get(cn) == en: hit += 1
            else: all_match = False
        if all_match and ecols: arr_ok += 1
    print(f'Walidacja AC — dokładność per-enum: {hit}/{tot} ({100*hit//tot}%) | '
          f'stany w pełni zgodne: {arr_ok}/{len(rows)}')
    return hit, tot

if __name__ == '__main__':
    print('== walidacja mapowania na wzorze AC ==')
    validate()
    print('== generacja ==')
    build()
